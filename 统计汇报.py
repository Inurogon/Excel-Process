#统计
import openpyxl
from openpyxl import load_workbook
import time

data=load_workbook('C:\\Users\\admin\\Downloads\\123.xlsx')
table=data.get_sheet_by_name('基础数据表')
PT_1=PT_2=PT_3=PT_4=XH_1=XH_2=XH_3=XH_4=ZB_1=ZB_2=ZB_3=ZB_4=CN_1=CN_2=CN_3=CN_4=Result_1=Result_2=0
for i in range(1,1577):
    Date=str(table.cell(row=i,column=18).value)
    Condition=str(table.cell(row=i,column=17).value)
    IN_T1=str(time.strftime('%Y%m%d',time.localtime(time.time())))
    Rs=str(table.cell(row=i,column=19).value)
    Area=str(table.cell(row=i,column=16).value)
    if Condition=="已安装":
        Result_1=Result_1+1
        if Date==IN_T1:
            Result_2=Result_2+1
            if Area=="普陀区":
                PT_1=PT_1+1
            elif Area=="徐汇区":
                XH_1=XH_1+1
            elif Area=="黄浦区":
                ZB_1=ZB_1+1
            elif Area=="长宁区":
                CN_1=CN_1+1
            else:
                continue;
            
    if Date==IN_T1 and Condition=="未安装":       
       if Area=="普陀区":
            PT_2=PT_2+1
       elif Area=="徐汇区":
            XH_2=XH_2+1
       elif Area=="黄浦区":
            ZB_2=ZB_2+1
       elif Area=="长宁区":
            CN_2=CN_2+1
       else:
            print(" ")
            
    if Date==IN_T1 and Rs=="无信号":
        if Area=="普陀区":
            PT_3=PT_3+1
        elif Area=="徐汇区":
            XH_3=XH_3+1
        elif Area=="黄浦区":
            ZB_3=ZB_3+1
        elif Area=="长宁区":
            CN_3=CN_3+1
        else:
            print(" ")
            
    if Date==IN_T1 and Rs=="4G Loading":
        if Area=="普陀区":
            PT_4=PT_4+1
        elif Area=="徐汇区":
            XH_4=XH_4+1
        elif Area=="黄浦区":
            ZB_4=ZB_4+1
        elif Area=="长宁区":
            CN_4=CN_4+1
        else:
            print(" ")

        
print('**********************************************\n')
print('*                                            *\n')
print('*           12月26日模块安装统计结果         *\n')
print('*                                            *\n')
print('**********************************************\n')
print('      已安装     未安装     无信号     4GLOAD    ')
print('普陀      '+str(PT_1)+ '        '+str(PT_2)+'           '+str(PT_3)+'          '+str(PT_4)+'  ')
print('徐汇      '+str(XH_1)+ '        '+str(XH_2)+'           '+str(XH_3)+'          '+str(XH_4)+'  ')
print('长宁     '+str(CN_1)+ '       '+str(CN_2)+'           '+str(CN_3)+'          '+str(CN_4)+'  ')
print('黄浦     '+str(ZB_1)+ '       '+str(ZB_2)+'           '+str(ZB_3)+'         '+str(ZB_4)+'  ')
print('***********************************************')
print('                            今日总计完成了'+str(Result_2)+'个 ')
print('黄浦区今日安装明细：')
#print('-----------------------------------------------')
print('张：13个已安装/10个未安装                       ')
print('秦：09个已安装/13个未安装                       ')                                                   
print('------------------------------------------------')
print('静安区完成2个未安装 并且与普陀区已全部完成      ')
print('\n')
print('✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡✡')
print('                                               ')
print(' 目前总计完成了站点安装'+str(Result_1)+'个,未安装554个,剩余68个    ')          
print('                                               ')
print('         若有数据错误，请及时与我联系          ')
print('                                               ')
print('###############################################')
print('                                               ')
print('         通知：不要忘记向我索取加班单          ')
print('                                               ')
print('                                               ')
print('                             Created by  ☠    ')
print('                            *此为程序自动生成  ')
