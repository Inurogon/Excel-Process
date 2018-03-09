#逆向查询
import openpyxl
from openpyxl import load_workbook

data=load_workbook('C:\\Users\\admin\\Downloads\\123.xlsx')
table_1=data.get_sheet_by_name('基础数据表')
table_2=data.get_sheet_by_name('基础数据表')
Input_Num=input("请输入站名：")
for i in range(2,1527):
    Stop_Name=str(table_1.cell(row=i,column=3).value)
    Stop_Name_NewSheet=str(table_2.cell(row=i,column=5).value)
    if Stop_Name==Input_Num and table_1.cell(row=i,column=17).value is None:
        Stop_Num=int(table_1.cell(row=i,column=2).value)
        Stop_Num_NewSheet=str(table_2.cell(row=i,column=2).value)
        print("逆向已匹配到编号为   "+str(Stop_Num)+":"+str(Input_Num)+"   (尚未使用)")
        print("逆向已匹配到55寸屏   "+str(Stop_Num_NewSheet)+":"+str(Input_Num))
        print("\n")
    else:
        continue;


    



                
