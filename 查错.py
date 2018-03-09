#查错
import openpyxl
from openpyxl import load_workbook

data=load_workbook('C:\\Users\\admin\\Downloads\\123.xlsx')
table=data.get_sheet_by_name('基础数据表')
print("正在执行查错程序,请稍后:")
for i in range(2,1577):
    ICC1=str(table.cell(row=i,column=6).value)
    Stop_Name_1=str(table.cell(row=i,column=3).value)
    Stop_Num_1=str(table.cell(row=i,column=2).value)
    Number_1=str(table.cell(row=i,column=1).value)
    for n in range(i+1,1577):
            ICC2=str(table.cell(row=i+1,column=6).value)
            Stop_Name_2=str(table.cell(row=i+1,column=3).value)
            Stop_Num_2=str(table.cell(row=i+1,column=2).value)
            Number_2=str(table.cell(row=i+1,column=1).value)
            if ICC1 == ICC2:
                    print('序号'+Number_1+Stop_Name_1+':'+Stop_Num_1+'  ICCO号出现错误:'+ICC1)
                    print('序号'+Number_2+Stop_Name_2+':'+Stop_Num_2+'  ICCO号出现错误:'+ICC2)
                    print('-------------------以上信息重复--------------------')
                    break;
            else:
                    continue;
                
                
             
    
    
