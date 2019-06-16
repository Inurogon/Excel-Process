#服务器日志打印#
import codecs
import os
import sys
import traceback
import win32con
import win32evtlog
import win32evtlogutil
import time
import winerror
import openpyxl
from openpyxl import Workbook
import re
import datetime
from openpyxl.styles import Border, Side, Font


def trans_format(time_string, from_format, to_format='%Y.%m.%d %H:%M:%S'):
    """
    @note 时间格式转化
    :param time_string:
    :param from_format:
    :param to_format:
    :return:
    """
    time_struct = time.strptime(time_string,from_format)
    times = time.strftime(to_format, time_struct)
    return times

def Get_SecureEvents(logtypes,basepath):
    server="localhost"
    for logtype in logtypes:
        path=os.path.join(basepath,server+'_'+logtype+'_log.log')
        getEventLogs(server,logtype,path)
  

def getEventLogs(server,logtype,logpath):
    print('正在获取'+logtype+'日志')
    log=codecs.open(logpath,encoding='utf-8',mode='w')
#读取本机的系统日志
    hand=win32evtlog.OpenEventLog(server,logtype)
#读取系统日志总行数
    total=win32evtlog.GetNumberOfEventLogRecords(hand)
#按序读取
    flags=win32evtlog.EVENTLOG_FORWARDS_READ|win32evtlog.EVENTLOG_SEQUENTIAL_READ
##错误级别字典
    evt_dict = {win32con.EVENTLOG_AUDIT_FAILURE: 'EVENTLOG_AUDIT_FAILURE',
                win32con.EVENTLOG_AUDIT_SUCCESS: 'EVENTLOG_AUDIT_SUCCESS',
                win32con.EVENTLOG_INFORMATION_TYPE: 'EVENTLOG_INFORMATION_TYPE',
                win32con.EVENTLOG_WARNING_TYPE: 'EVENTLOG_WARNING_TYPE',
                win32con.EVENTLOG_ERROR_TYPE: 'EVENTLOG_ERROR_TYPE'}

    try:
        events=1
        j=2
        WaitTime=datetime.datetime.now()
        WaitTime=(datetime.datetime.now()+datetime.timedelta(minutes=-1)).strftime("%Y%m%d%H%M%S")
        while events:
                events=win32evtlog.ReadEventLog(hand,flags,0)
                for ev_obj in events:
                    the_time=ev_obj.TimeGenerated.Format()
                    the_time=trans_format(the_time,'%a %b %d %H:%M:%S %Y','%Y%m%d%H%M%S')
                    if the_time<WaitTime:
                        continue
                    else:
                        evt_id=winerror.HRESULT_CODE(ev_obj.EventID)
                        computer=ev_obj.ComputerName
                        cat=ev_obj.EventCategory
                        record=ev_obj.RecordNumber
                        msg=win32evtlogutil.SafeFormatMessage(ev_obj,logtype)
                        source=str(ev_obj.SourceName)
                        if not ev_obj.EventType in evt_dict.keys():
                            evt_type="unknow"
                        else:
                            evt_type=evt_dict[ev_obj.EventType]
                            log.write("Event Date/Time: %s\n" %the_time)
                            log.write("EventID /Type: %s / %s\n" %(evt_id,evt_type))
                            log.write("Record #%s\n" %record)
                            log.write("Source: %s\n\n" %source)
                            log.write(msg)
                            log.write('__________________________________________________')
                            log.write('\n\n')
                            AccountName=re.search('帐户名:\t\t.+',msg)
                            if AccountName==None:
                                continue
                            else:
                                AccountName=AccountName.group(0)
                                AccountName=AccountName[6:]
                                DateType=re.search('对象类型:\t\t.+',msg)  #\s也可以匹配上
                                if DateType==None:
                                    continue          
                                else:
                                    DateType=DateType.group(0)
                                    DateType=DateType[7:]
                                    DateType=DateType[0:-1]
                                    if DateType=='File':
                                        Result=re.search("访问:\t\t.+",msg)
                                        if Result==None:
                                            continue
                                        else:
                                            Result=Result.group(0)
                                            Result=Result[7:]
                                            Result=Result[0:-1]
                                            if Result=='1537': 
                                               print('删除')
                                               DoSomething=re.search('对象名:\t\t.+',msg)
                                               DoSomething=DoSomething.group(0)
                                               ws.cell(row=j,column=1).value=the_time
                                               ws.cell(row=j,column=2).value=AccountName
                                               ws.cell(row=j,column=3).value=DoSomething
                                               ws.cell(row=j,column=4).value='删除了该文件'
                                               j=j+1
                                            elif Result=='4416':
                                                 print('读取')
                                                 DoSomething=re.search('对象名:\t\t.+',msg)
                                                 DoSomething=DoSomething.group(0)
                                                 ws.cell(row=j,column=1).value=the_time
                                                 ws.cell(row=j,column=2).value=AccountName
                                                 ws.cell(row=j,column=3).value=DoSomething
                                                 ws.cell(row=j,column=4).value='读取该文件'
                                                 j=j+1
                                            elif Result=='4417':
                                                 print('写入')
                                                 DoSomething=re.search('对象名:\t\t.+',msg)
                                                 DoSomething=DoSomething.group(0)
                                                 ws.cell(row=j,column=1).value=the_time
                                                 ws.cell(row=j,column=2).value=AccountName
                                                 ws.cell(row=j,column=3).value=DoSomething
                                                 ws.cell(row=j,column=4).value='写入了文件'
                                                 j=j+1
                                            else:
                                               continue
                                           #print('WIRTEDATA OR WIRTELIST')
                                           #Dir=re.search("对象名:\t\t.+",msg)
                
                
        
        
                    
    except:
            print(traceback.print_exc(sys.exc_info()))


    

if __name__=="__main__":
    logtypes=["Security"]
    wb=Workbook()
    ws=wb.active
    ws.title='日志管理'
    border = Border(left=Side(style='medium',color='FF3030'),right=Side(style='medium',color='FF3030'),top=Side(style='medium',color='FF3030'),bottom=Side(style='medium',color='FF3030'),diagonal=Side(style='medium',color='FF3030'),diagonal_direction=0,outline=Side(style='medium',color='FF3030'),vertical=Side(style='medium',color='FF3030'),horizontal=Side(style='medium',color='FF3030'))
    ws.cell(row=1,column=1).value='事件时间'
    ws.cell(row=1,column=2).value='账户名称'
    ws.cell(row=1,column=3).value='文件路径'
    ws.cell(row=1,column=4).value='操作事项'
    ws['A1'].border=border
    ws['B1'].border=border
    ws['C1'].border=border
    ws['D1'].border=border 
    for p in range(10000):
        Get_SecureEvents(logtypes,'D:\\')
        print('已循环')
        wb.save('D:\\1.xlsx')
        time.sleep(60)

        
