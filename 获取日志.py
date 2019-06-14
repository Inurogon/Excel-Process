#服务器日志打印#
import codecs
import os
import sys
import traceback
import win32con
import win32evtlog
import win32evtlogutil
import time
import winerror
import openpyxl
from openpyxl import Workbook
import re
import datetime



def Get_SecureEvents(logtypes,basepath):
    server="localhost"
    for logtype in logtypes:
        path=os.path.join(basepath,server+'_'+logtype+'_log.log')
        getEventLogs(server,logtype,path)
  

def getEventLogs(server,logtype,logpath):
    print('正在获取'+logtype+'日志')
    log=codecs.open(logpath,encoding='utf-8',mode='w')
#读取本机的系统日志
    hand=win32evtlog.OpenEventLog(server,logtype)
#读取系统日志总行数
    total=win32evtlog.GetNumberOfEventLogRecords(hand)
#按序读取
    flags=win32evtlog.EVENTLOG_FORWARDS_READ|win32evtlog.EVENTLOG_SEQUENTIAL_READ
##错误级别字典
    evt_dict = {win32con.EVENTLOG_AUDIT_FAILURE: 'EVENTLOG_AUDIT_FAILURE',
                win32con.EVENTLOG_AUDIT_SUCCESS: 'EVENTLOG_AUDIT_SUCCESS',
                win32con.EVENTLOG_INFORMATION_TYPE: 'EVENTLOG_INFORMATION_TYPE',
                win32con.EVENTLOG_WARNING_TYPE: 'EVENTLOG_WARNING_TYPE',
                win32con.EVENTLOG_ERROR_TYPE: 'EVENTLOG_ERROR_TYPE'}

    try:
        events=1
        i=1
        j=2
        WaitTime='None'
        for i in range(10000):
            while events:
                events=win32evtlog.ReadEventLog(hand,flags,0)
                for ev_obj in events:
                    the_time=ev_obj.TimeGenerated.Format() #'12/23/999 15:54:09'
                    if WaitTime=='None':
                        evt_id=winerror.HRESULT_CODE(ev_obj.EventID)
                        computer=ev_obj.ComputerName
                        cat=ev_obj.EventCategory
                        record=ev_obj.RecordNumber
                        msg=win32evtlogutil.SafeFormatMessage(ev_obj,logtype)
                        source=str(ev_obj.SourceName)
                        if not ev_obj.EventType in evt_dict.keys():
                            evt_type="unknow"
                        else:
                            evt_type=evt_dict[ev_obj.EventType]
                            log.write("Event Date/Time: %s\n" %the_time)
                            log.write("EventID /Type: %s / %s\n" %(evt_id,evt_type))
                            log.write("Record #%s\n" %record)
                            log.write("Source: %s\n\n" %source)
                            log.write(msg)
                            AccountName=re.search('帐户名:\t\t.+',msg)
                            if AccountName==None:
                                pass
                            else:
                                AccountName=AccountName.group(0)
                                AccountName=AccountName[6:]
                                DateType=re.search('对象类型:\t\t.+',msg)  #\s也可以匹配上
                                if DateType==None:
                                    pass          
                                else:
                                    DateType=DateType.group(0)
                                    DateType=DateType[7:]
                                    DateType=DateType[0:-1]
                                    if DateType=='File':
                                        Result=re.search("访问:\t\t.+",msg)
                                        if Result==None:
                                            pass
                                        else:
                                            Result=Result.group(0)
                                            Result=Result[7:]
                                            Result=Result[0:-1]
                                            if Result=='1537':
                                               print('删除')
                                               DoSomething=re.search('对象名:\t\t.+',msg)
                                               DoSomething=DoSomething.group(0)
                                               ws.cell(row=j,column=1).value=the_time
                                               ws.cell(row=j,column=2).value=AccountName
                                               ws.cell(row=j,column=3).value=DoSomething
                                               ws.cell(row=j,column=4).value='删除了该文件'
                                               j=j+1
                                            elif Result=='4416':
                                                 print('读取')
                                                 DoSomething=re.search('对象名:\t\t.+',msg)
                                                 DoSomething=DoSomething.group(0)
                                                 ws.cell(row=j,column=1).value=the_time
                                                 ws.cell(row=j,column=2).value=AccountName
                                                 ws.cell(row=j,column=3).value=DoSomething
                                                 ws.cell(row=j,column=4).value='读取该文件'
                                                 j=j+1
                                                 WaitTime=datetime.datetime.now()
                                                 #print(time.strptime(WaitTime,'%Y/%m/%d'))
                                                 #print(int(time.strptime(WaitTime,'%Y/%m/%d')))
                                                 time.sleep(1)

                    elif int(time.strptime(WaitTime,'%Y/%m/%d'))>=int(time.strptime(the_time,'%Y/%m/%d')):
                        pass
                    else:
                        evt_id=winerror.HRESULT_CODE(ev_obj.EventID)
                        computer=ev_obj.ComputerName
                        cat=ev_obj.EventCategory
                        record=ev_obj.RecordNumber
                        msg=win32evtlogutil.SafeFormatMessage(ev_obj,logtype)
                        source=str(ev_obj.SourceName)
                        if not ev_obj.EventType in evt_dict.keys():
                            evt_type="unknow"
                        else:
                            evt_type=evt_dict[ev_obj.EventType]
                            log.write("Event Date/Time: %s\n" %the_time)
                            log.write("EventID /Type: %s / %s\n" %(evt_id,evt_type))
                            log.write("Record #%s\n" %record)
                            log.write("Source: %s\n\n" %source)
                            log.write(msg)
                            AccountName=re.search('帐户名:\t\t.+',msg)
                            if AccountName==None:
                                pass
                            else:
                                AccountName=AccountName.group(0)
                                AccountName=AccountName[6:]
                                DateType=re.search('对象类型:\t\t.+',msg)  #\s也可以匹配上
                                if DateType==None:
                                    pass          
                                else:
                                    DateType=DateType.group(0)
                                    DateType=DateType[7:]
                                    DateType=DateType[0:-1]
                                    if DateType=='File':
                                        Result=re.search("访问:\t\t.+",msg)
                                        if Result==None:
                                            pass
                                        else:
                                            Result=Result.group(0)
                                            Result=Result[7:]
                                            Result=Result[0:-1]
                                            if Result=='1537':
                                               print('删除')
                                               DoSomething=re.search('对象名:\t\t.+',msg)
                                               DoSomething=DoSomething.group(0)
                                               ws.cell(row=j,column=1).value=the_time
                                               ws.cell(row=j,column=2).value=AccountName
                                               ws.cell(row=j,column=3).value=DoSomething
                                               ws.cell(row=j,column=4).value='删除了该文件'
                                               j=j+1
                                            elif Result=='4416':
                                                 print('读取')
                                                 DoSomething=re.search('对象名:\t\t.+',msg)
                                                 DoSomething=DoSomething.group(0)
                                                 ws.cell(row=j,column=1).value=the_time
                                                 ws.cell(row=j,column=2).value=AccountName
                                                 ws.cell(row=j,column=3).value=DoSomething
                                                 ws.cell(row=j,column=4).value='读取该文件'
                                                 j=j+1
                                            else:
                                               pass
                                           #print('WIRTEDATA OR WIRTELIST')
                                           #Dir=re.search("对象名:\t\t.+",msg)

                            log.write('__________________________________________________')
                            log.write('\n\n')
                            i=i+1
                #WaitTime=datetime.datetime.now()
                #time.sleep(1)
        
        
                    
    except:
            print(traceback.print_exc(sys.exc_info()))


    

if __name__=="__main__":
    logtypes=["Security"]
    wb=Workbook()
    ws=wb.active
    ws.title='日志管理'
    Get_SecureEvents(logtypes,'D:\\')
    wb.save('D:\\1.xlsx')
    print('完成')
        


