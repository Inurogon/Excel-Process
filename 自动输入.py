#自动输入
import openpyxl
from openpyxl import load_workbook
import time

data=load_workbook('C:\\Users\\admin\Downloads\\123.xlsx')
table=data.get_sheet_by_name('基础数据表')
for k in range(0,100):
    Input_Num=input("请输入站号：")
    for i in range(2,1536):
        Stop_Num=int(table.cell(row=i,column=2).value)
        if int(Input_Num) == Stop_Num:
            Stop_Name=str(table.cell(row=i,column=3).value)
            print("匹配成功，该站信息为:"+Stop_Name)
            Result_1=input("请确认该信息是否正确，正确请按y，不正确请按任意键:")
            if Result_1 == "y":
                justice=input("请问是安装成功的点么:y.已安装/其他.未安装")
                if justice == "y":
                    ICC=str(input("请输入ICC ID后三至五位："))
                    if len(ICC) == 4:
                        ICC2=str("8986001300100050"+ICC)
                    elif len(ICC) == 3:
                        ICC2=str("89860013001000500"+ICC)
                    else:
                        ICC2=str("898600130010005"+ICC)
                    SN_Num=str(input("请输入完整的SN号："))
                    table.cell(row=i,column=6).value=ICC2
                    table.cell(row=i,column=5).value=int(ICC.replace("A","0"))/10-246+19801870000
                    table.cell(row=i,column=7).value=SN_Num
                    table.cell(row=i,column=17).value='已安装'
                    IN_T1=str(time.strftime('%Y%m%d',time.localtime(time.time())))
                    table.cell(row=i,column=18).value=IN_T1
                    table.cell(row=i,column=12).value='是'
                    Reason1=str(input("请输入备注信息："))
                    table.cell(row=i,column=19).value=Reason1
                    data.save('123.xlsx')
                    print("录入完成")
                    print("\n")
                    print("\n")
                    print("--------------------------------------------------")
                else:
                    Reason2=str(input("请输入失败的原因："))
                    table.cell(row=i,column=17).value='未安装'
                    IN_T2=str(time.strftime('%Y%m%d',time.localtime(time.time())))
                    table.cell(row=i,column=18).value=IN_T2
                    table.cell(row=i,column=19).value=Reason2
                    if Reason2==str('无信号'):
                        table.cell(row=i,column=12).value='否'
                    else:
                        table.cell(row=i,column=12).value='是'
                    data.save('123.xlsx')
                    print("录入完成")
                    print("\n")
                    print("\n")
                    print("--------------------------------------------------")
                
            else:
                
                
                break;
                
           
    else:
        continue;
   

